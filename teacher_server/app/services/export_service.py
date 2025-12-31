"""
数据导出服务
支持导出为Excel、CSV、JSON格式
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
import csv
import json
from app.models import StudentGroup, ChatMessage

class DataExportService:
    """数据导出服务类"""
    
    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        self.header_font = Font(bold=True, color='FFFFFF', size=11)
    
    def export_to_excel(self, groups, output_path=None):
        """导出为Excel格式（推荐）"""
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 1. 数据统计摘要
        self._create_summary_sheet(wb, groups)
        
        # 2. 学生基本信息
        self._create_student_info_sheet(wb, groups)
        
        # 3. 任务一数据
        self._create_task1_sheet(wb, groups)
        
        # 4. 任务二数据
        self._create_task2_sheet(wb, groups)
        
        # 5. 思考题数据
        self._create_thinking_sheet(wb, groups)
        
        # 6. 茶助教问答记录
        self._create_chat_sheet(wb, groups)
        
        # 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'茶文化课程数据_{timestamp}.xlsx'
        
        wb.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, wb, groups):
        """创建数据统计摘要工作表"""
        ws = wb.create_sheet('数据统计摘要', 0)  # 放在第一个位置
        
        # 标题
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = '茶文化课程数据统计摘要'
        title_cell.font = Font(bold=True, size=16, color='2E7D32')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # 导出信息
        ws['A3'] = '导出时间：'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = '数据总数：'
        ws['B4'] = len(groups)
        ws['A3'].font = Font(bold=True)
        ws['A4'].font = Font(bold=True)
        
        # 按班级统计
        ws['A7'] = '按班级统计'
        ws['A7'].font = Font(bold=True, size=14, color='2E7D32')
        
        headers = ['学校', '年级', '班级', '提交组数', '任务一完成', '任务二完成', '平均提问数']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(8, col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 按班级分组统计
        class_stats = {}
        for group in groups:
            key = (group.school, group.grade, group.class_number)
            if key not in class_stats:
                class_stats[key] = {
                    'count': 0,
                    'task1': 0,
                    'task2': 0,
                    'questions': 0
                }
            
            class_stats[key]['count'] += 1
            if group.task1:
                class_stats[key]['task1'] += 1
            if group.task2:
                class_stats[key]['task2'] += 1
            class_stats[key]['questions'] += len([m for m in group.chat_messages if m.role == 'user'])
        
        row_num = 9
        for (school, grade, class_num), stats in sorted(class_stats.items()):
            avg_questions = stats['questions'] / stats['count'] if stats['count'] > 0 else 0
            ws.cell(row_num, 1).value = school
            ws.cell(row_num, 2).value = grade
            ws.cell(row_num, 3).value = class_num
            ws.cell(row_num, 4).value = stats['count']
            ws.cell(row_num, 5).value = stats['task1']
            ws.cell(row_num, 6).value = stats['task2']
            ws.cell(row_num, 7).value = round(avg_questions, 1)
            
            # 设置边框
            for col in range(1, 8):
                ws.cell(row_num, col).border = self.thin_border
                ws.cell(row_num, col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
    
    def _create_student_info_sheet(self, wb, groups):
        """创建学生基本信息工作表"""
        ws = wb.create_sheet('学生基本信息')
        
        # 表头
        headers = [
            '小组编号', '学校', '年级', '班级', '活动日期', '成员人数', 
            '成员姓名', '提交时间', '任务一完成', '任务二完成', 
            '思考题一完成', '思考题二完成', '创意题完成', '茶助教提问数'
        ]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据行
        for group in groups:
            member_names = ', '.join([m.member_name for m in group.members])
            
            # 统计完成情况
            task1_done = '是' if group.task1 else '否'
            task2_done = '是' if group.task2 else '否'
            thinking1_done = '是' if any(t.question_type == 'thinking1' for t in group.thinking_questions) else '否'
            thinking2_done = '是' if any(t.question_type == 'thinking2' for t in group.thinking_questions) else '否'
            creative_done = '是' if any(t.question_type == 'creative' for t in group.thinking_questions) else '否'
            
            # 茶助教提问数
            chat_count = len([m for m in group.chat_messages if m.role == 'user'])
            
            row = [
                group.group_number or '',
                group.school,
                group.grade,
                group.class_number,
                group.activity_date.strftime('%Y-%m-%d') if group.activity_date else '',
                group.member_count,
                member_names,
                group.submit_time.strftime('%Y-%m-%d %H:%M:%S') if group.submit_time else '',
                task1_done,
                task2_done,
                thinking1_done,
                thinking2_done,
                creative_done,
                chat_count
            ]
            ws.append(row)
            
            # 设置数据行样式
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
    
    def _create_task1_sheet(self, wb, groups):
        """创建任务一数据工作表"""
        ws = wb.create_sheet('任务一-泡茶体验')
        
        headers = [
            '小组编号', '学校', '年级', '班级', '茶品名', '老师茶品名', '茶类',
            '水温', '冲泡时长', 
            '干茶-色泽', '干茶-香气', '干茶-形状', '干茶-滋味',
            '茶汤-色泽', '茶汤-香气', '茶汤-形状', '茶汤-滋味',
            '叶底-色泽', '叶底-香气', '叶底-形状', '叶底-滋味',
            '思考题答案', '答案字数'
        ]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据行
        for group in groups:
            if not group.task1:
                continue
            
            task1 = group.task1
            records = task1.get_sensory_records()
            
            row = [
                group.group_number or '',
                group.school,
                group.grade,
                group.class_number,
                task1.tea_name or '',
                task1.teacher_tea_name or '',
                task1.tea_category or '',
                task1.water_temperature or '',
                task1.brewing_duration or '',
                records.get('dryTea', {}).get('color', ''),
                records.get('dryTea', {}).get('aroma', ''),
                records.get('dryTea', {}).get('shape', ''),
                records.get('dryTea', {}).get('taste', ''),
                records.get('teaLiquor', {}).get('color', ''),
                records.get('teaLiquor', {}).get('aroma', ''),
                records.get('teaLiquor', {}).get('shape', ''),
                records.get('teaLiquor', {}).get('taste', ''),
                records.get('spentLeaves', {}).get('color', ''),
                records.get('spentLeaves', {}).get('aroma', ''),
                records.get('spentLeaves', {}).get('shape', ''),
                records.get('spentLeaves', {}).get('taste', ''),
                task1.reflection_answer or '',
                len(task1.reflection_answer) if task1.reflection_answer else 0
            ]
            ws.append(row)
            
            # 设置数据行样式
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        for col in range(1, 22):
            ws.column_dimensions[chr(64 + col)].width = 15
        ws.column_dimensions['V'].width = 50  # 思考题答案列
    
    def _create_task2_sheet(self, wb, groups):
        """创建任务二数据工作表"""
        ws = wb.create_sheet('任务二-泡出心中茶')
        
        headers = [
            '小组编号', '学校', '年级', '班级', '茶品名', '水温', '出汤时长',
            '茶汤-色泽', '茶汤-香气', '茶汤-滋味',
            '符合预期', '不符合预期', '思考题答案', '答案字数'
        ]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据行
        for group in groups:
            if not group.task2:
                continue
            
            task2 = group.task2
            
            row = [
                group.group_number or '',
                group.school,
                group.grade,
                group.class_number,
                task2.tea_name or '',
                task2.water_temperature or '',
                task2.steeping_duration or '',
                task2.tea_color or '',
                task2.tea_aroma or '',
                task2.tea_taste or '',
                '是' if task2.meets_expectation else '否',
                '是' if task2.not_meets_expectation else '否',
                task2.reflection_answer or '',
                len(task2.reflection_answer) if task2.reflection_answer else 0
            ]
            ws.append(row)
            
            # 设置数据行样式
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        ws.column_dimensions['M'].width = 50  # 思考题答案列
    
    def _create_thinking_sheet(self, wb, groups):
        """创建思考题数据工作表"""
        ws = wb.create_sheet('思考题')
        
        headers = [
            '小组编号', '学校', '年级', '班级',
            '思考题一答案', '思考题一字数',
            '思考题二答案', '思考题二字数',
            '创意题答案', '创意题字数',
            '总字数'
        ]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据行
        for group in groups:
            thinking1 = next((t for t in group.thinking_questions if t.question_type == 'thinking1'), None)
            thinking2 = next((t for t in group.thinking_questions if t.question_type == 'thinking2'), None)
            creative = next((t for t in group.thinking_questions if t.question_type == 'creative'), None)
            
            thinking1_answer = thinking1.answer if thinking1 else ''
            thinking2_answer = thinking2.answer if thinking2 else ''
            creative_answer = creative.answer if creative else ''
            
            thinking1_len = len(thinking1_answer)
            thinking2_len = len(thinking2_answer)
            creative_len = len(creative_answer)
            total_len = thinking1_len + thinking2_len + creative_len
            
            row = [
                group.group_number or '',
                group.school,
                group.grade,
                group.class_number,
                thinking1_answer,
                thinking1_len,
                thinking2_answer,
                thinking2_len,
                creative_answer,
                creative_len,
                total_len
            ]
            ws.append(row)
            
            # 设置数据行样式
            row_num = ws.max_row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row_num, col)
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 50
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
    
    def _create_chat_sheet(self, wb, groups):
        """创建茶助教问答记录工作表"""
        ws = wb.create_sheet('茶助教问答记录')
        
        headers = [
            '小组编号', '学校', '年级', '班级', '对话序号', 
            '角色', '消息内容', '消息长度', '提交时间'
        ]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = self.header_font
            cell.fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
        
        # 数据行
        for group in groups:
            chat_messages = sorted(group.chat_messages, key=lambda x: x.message_index)
            
            for msg in chat_messages:
                role_name = '学生' if msg.role == 'user' else '茶助教'
                
                row = [
                    group.group_number or '',
                    group.school,
                    group.grade,
                    group.class_number,
                    msg.message_index + 1,
                    role_name,
                    msg.content,
                    len(msg.content),
                    msg.submit_time.strftime('%Y-%m-%d %H:%M:%S') if msg.submit_time else ''
                ]
                ws.append(row)
                
                # 设置数据行样式
                row_num = ws.max_row
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row_num, col)
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 80  # 消息内容列
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
    
    def export_to_json(self, groups, output_path=None):
        """导出为JSON格式（适合AI分析）"""
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'茶文化课程数据_{timestamp}.json'
        
        data = {
            'export_time': datetime.now().isoformat(),
            'total_groups': len(groups),
            'groups': []
        }
        
        for group in groups:
            group_data = {
                'group_info': {
                    'group_number': group.group_number,
                    'school': group.school,
                    'grade': group.grade,
                    'class_number': group.class_number,
                    'activity_date': group.activity_date.isoformat() if group.activity_date else None,
                    'submit_time': group.submit_time.isoformat() if group.submit_time else None,
                    'members': [m.member_name for m in group.members]
                },
                'task1': group.task1.to_dict() if group.task1 else None,
                'task2': group.task2.to_dict() if group.task2 else None,
                'thinking_questions': {
                    t.question_type: t.answer
                    for t in group.thinking_questions
                },
                'chat_messages': [
                    {
                        'index': msg.message_index,
                        'role': msg.role,
                        'content': msg.content
                    }
                    for msg in sorted(group.chat_messages, key=lambda x: x.message_index)
                ],
                'statistics': {
                    'student_questions': len([m for m in group.chat_messages if m.role == 'user']),
                    'ai_responses': len([m for m in group.chat_messages if m.role == 'assistant']),
                    'task1_char_count': group.get_task1_char_count(),
                    'task2_char_count': group.get_task2_char_count()
                }
            }
            data['groups'].append(group_data)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return output_path

